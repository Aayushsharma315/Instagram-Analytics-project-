"""
03_build_excel_workbook.py
------------------------------------------------
Builds the Excel deliverable (excel/instagram_analytics.xlsx):
  - "Raw Data" sheet: full cleaned dataset
  - "Summary" sheet: formula-driven KPIs and breakdown tables
    (SUMIFS / AVERAGEIFS / COUNTIFS referencing the Raw Data sheet,
     so the summary recalculates if the raw data changes)

Run: python scripts/03_build_excel_workbook.py
Then: python scripts/office/recalc.py excel/instagram_analytics.xlsx
(recalc script comes from the xlsx skill toolkit)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

df = pd.read_csv("data/processed/instagram_posts_clean.csv")

wb = Workbook()

# ---------------------------------------------------------------
# Sheet 1: Raw Data
# ---------------------------------------------------------------
ws_raw = wb.active
ws_raw.title = "Raw Data"

headers = list(df.columns)
for col_idx, h in enumerate(headers, start=1):
    cell = ws_raw.cell(row=1, column=col_idx, value=h)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")

for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        ws_raw.cell(row=row_idx, column=col_idx, value=value).font = Font(name="Arial", size=10)

for col_idx, h in enumerate(headers, start=1):
    ws_raw.column_dimensions[get_column_letter(col_idx)].width = max(12, len(h) + 2)

n_rows = len(df)  # data rows, header is row 1, data is rows 2..n_rows+1

# Column letters we need for formulas
col_letter = {name: get_column_letter(i + 1) for i, name in enumerate(headers)}

# ---------------------------------------------------------------
# Sheet 2: Summary (formula-driven)
# ---------------------------------------------------------------
ws = wb.create_sheet("Summary")
title_font = Font(name="Arial", bold=True, size=14)
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")
label_font = Font(name="Arial", bold=True)
data_font = Font(name="Arial")

ws["A1"] = "Instagram Engagement Analytics — Summary"
ws["A1"].font = title_font
ws.merge_cells("A1:D1")

# --- KPI block ---
ws["A3"] = "Key Metrics"
ws["A3"].font = label_font

kpis = [
    ("Total Posts", f"=COUNTA('Raw Data'!{col_letter['post_id']}2:{col_letter['post_id']}{n_rows+1})"),
    ("Total Reach", f"=SUM('Raw Data'!{col_letter['reach']}2:{col_letter['reach']}{n_rows+1})"),
    ("Total Engagement", f"=SUM('Raw Data'!{col_letter['total_engagement']}2:{col_letter['total_engagement']}{n_rows+1})"),
    ("Average Engagement Rate", f"=AVERAGE('Raw Data'!{col_letter['engagement_rate']}2:{col_letter['engagement_rate']}{n_rows+1})"),
    ("Total Follows Gained", f"=SUM('Raw Data'!{col_letter['follows_gained']}2:{col_letter['follows_gained']}{n_rows+1})"),
]

row = 4
for label, formula in kpis:
    ws.cell(row=row, column=1, value=label).font = data_font
    fcell = ws.cell(row=row, column=2, value=formula)
    fcell.font = data_font
    if "Rate" in label:
        fcell.number_format = "0.0%"
    else:
        fcell.number_format = "#,##0"
    row += 1

# --- Breakdown by Post Type ---
start_row = row + 2
ws.cell(row=start_row, column=1, value="Performance by Post Type").font = label_font
start_row += 1

headers_bt = ["Post Type", "Post Count", "Avg Engagement Rate", "Avg Reach"]
for i, h in enumerate(headers_bt):
    c = ws.cell(row=start_row, column=1 + i, value=h)
    c.font = header_font
    c.fill = header_fill

post_types = sorted(df["post_type"].unique())
data_range_type = f"'Raw Data'!{col_letter['post_type']}2:{col_letter['post_type']}{n_rows+1}"
data_range_er = f"'Raw Data'!{col_letter['engagement_rate']}2:{col_letter['engagement_rate']}{n_rows+1}"
data_range_reach = f"'Raw Data'!{col_letter['reach']}2:{col_letter['reach']}{n_rows+1}"

for i, pt in enumerate(post_types):
    r = start_row + 1 + i
    ws.cell(row=r, column=1, value=pt).font = data_font
    ws.cell(row=r, column=2, value=f'=COUNTIF({data_range_type},A{r})').font = data_font
    er_cell = ws.cell(row=r, column=3, value=f'=AVERAGEIF({data_range_type},A{r},{data_range_er})')
    er_cell.font = data_font
    er_cell.number_format = "0.0%"
    reach_cell = ws.cell(row=r, column=4, value=f'=AVERAGEIF({data_range_type},A{r},{data_range_reach})')
    reach_cell.font = data_font
    reach_cell.number_format = "#,##0"

# --- Breakdown by Day of Week ---
start_row2 = start_row + len(post_types) + 3
ws.cell(row=start_row2, column=1, value="Performance by Day of Week").font = label_font
start_row2 += 1

for i, h in enumerate(["Day", "Post Count", "Avg Engagement Rate"]):
    c = ws.cell(row=start_row2, column=1 + i, value=h)
    c.font = header_font
    c.fill = header_fill

days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
data_range_day = f"'Raw Data'!{col_letter['day_of_week']}2:{col_letter['day_of_week']}{n_rows+1}"

for i, day in enumerate(days):
    r = start_row2 + 1 + i
    ws.cell(row=r, column=1, value=day).font = data_font
    ws.cell(row=r, column=2, value=f'=COUNTIF({data_range_day},A{r})').font = data_font
    er_cell = ws.cell(row=r, column=3, value=f'=AVERAGEIF({data_range_day},A{r},{data_range_er})')
    er_cell.font = data_font
    er_cell.number_format = "0.0%"

for col in ["A", "B", "C", "D"]:
    ws.column_dimensions[col].width = 22

wb.save("excel/instagram_analytics.xlsx")
print("Workbook saved -> excel/instagram_analytics.xlsx")
print("IMPORTANT: run recalc.py on this file before opening, so formulas show cached values.")
